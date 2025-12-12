# 10.07.2023 validation code for repeated choice dataset
# Yujia Yang

# install package first if not exist in current environment
import importlib
import subprocess
import sys
import math
import datetime
from pathlib import Path


def ensure_package(package_name, import_name=None):
    import_name = import_name or package_name
    try:
        importlib.import_module(import_name)
    except ImportError:
        print(f"Package '{package_name}' not found. Installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])

# import glob
ensure_package('pandas')
ensure_package('numpy')
ensure_package('scipy')
ensure_package('openpyxl')

import pandas as pd
import numpy as np
from scipy.stats import binomtest
from scipy.stats import chisquare
from openpyxl import Workbook, load_workbook


def get_names(name_pattern):
    current_dir = Path.cwd()  # get current working direction
    file_names = list(current_dir.glob(name_pattern))  # get matched file name list
    return file_names


# get current path
current_file = Path(__file__)
validation_dir = current_file.parent

# get parent path
project_root = validation_dir.parent

# get data folder path
data_dir = project_root / 'data'

# save path and name
save_name = 'stationary.xlsx'
save_path = validation_dir / save_name

# get all data sub-folders' name
file_list = [f.name for f in data_dir.iterdir() if f.is_dir()]
file_list = sorted(file_list)
validation_df = pd.DataFrame(columns=['study', 'consistent_outcomes', 'consistent_probabilities',
                                      'consistent_probabilities_problem', 'high_ev_choices', 'high_ev_choices_problems',
                                      'high_ev_choices_problems_detail', 'high_ev_choices_prob_per_problem',
                                      'high_mean_choices', 'high_mean_choices_problems', 'high_mean_choices_problems_detail',
                                      'high_mean_choices_prob_per_problem', 'equal_mean_choices', 'equal_mean_choices_prob_per_problem'])

# read file with pandas
info_dir = data_dir / 'feature_table.csv'
df = pd.read_csv(info_dir)
# filter data
paradigms = ['social binary prediction', 'binary prediction', 'description bandits', 'lottery bandits', 'observe or bet', 'probabilistic feedback binary prediction', 'probability learning', 'other', 'free sampling', 'regulated sampling']
filtered_df = df[df['paradigm'].isin(paradigms)]
studies_list = filtered_df['study'].str.split('_').str[0]
paper_list = studies_list.unique()
paper_list = [x for x in paper_list if x in file_list]
paper_list = sorted(paper_list)
# some analyzing variables
alpha = 0.01  # statistical parameters
# create empty file
wb = Workbook()
ws = wb.active
ws.append(validation_df.columns.to_list())
wb.save(save_path)
# start time
current_time = datetime.datetime.now()
print('start: ', current_time)
# read paper list
for folder in paper_list:  # file_list:
    target_path = data_dir / folder
    processed_path = target_path / 'processed'
    data_name_list = list(processed_path.glob('*_data.csv'))
    # read studies
    for data_path in data_name_list:
        study = folder + data_path.name[-11:-9]
        both = 1  # one problem have how many kinds of feedback type
        print(study)
        data = pd.read_csv(data_path)
        cleaned_data = pd.DataFrame()  # for data exclude trials has inconsistent values
        high_ev_trial_num = 0  # trial number of problems with high ev
        option_file = data_path.with_name(data_path.name[:-9] + '_options.csv')
        option_inf = pd.read_csv(option_file)
        option_columns = option_inf.columns.tolist()
        # skip datasets with continues/dynamic distribution
        if len(option_columns) == 2:
            print('continues/dynamic distribution')
            continue
        stages = data['stage'].unique()
        if len(stages) == 2:
            data = data[data['stage'] == 2]
        # filter no choice trials
        data = data[data['choice'] != '']
        # filter no outcome trials
        data = data[data['outcome'] != '']
        data = data.dropna(subset=['choice', 'outcome'])
        data = data[~pd.to_numeric(data['outcome'], errors='coerce').notna()]
        if len(data) == 0:
            print('no valuable outcome')
            continue
        data = data[data['outcome'].str.contains(':', case=False, na=False)].reset_index()  # filter only choice or only feedback data
        # skip no numerical outcome data
        value = data.loc[0, 'outcome']
        if not value[-1].isdigit():
            print('no numerical outcome')
            continue
        # validation variables for each study
        consistent_outcomes_num = 0
        consistent_probabilities = []
        consistent_probabilities_problem = 0
        high_ev_choices_num = []
        high_ev_choices_problems_num = []
        high_ev_choices_problem_prob = []
        high_mean_choices_num = []
        high_mean_choices_problems_num = []
        high_mean_choices_problem_prob = []
        equal_mean_choices_num = []
        equal_mean_choices_problem_prob = []
        experienced_trials_all = 0
        ignore_trials_num = 0
        participants_num = 0
        # for each problem
        option_pairs = data.groupby(data['options'].apply(lambda x: tuple(sorted(x.split('_')))))
        for option_pair in option_pairs:
            data_problem = option_pair[1]
            option_number = data_problem['choice'].nunique()
            # split outcome
            # partial feedback
            data_partial = data_problem[~data_problem['outcome'].str.contains('_', case=False)]
            if not data_partial.empty:
                # validation_variables each problem
                consistent_outcomes_p = 0
                consistent_probabilities_p = []
                consistent_probabilities_problem_p = 0
                high_ev_option_value_p = 0
                high_ev_option_p = 0
                high_ev_choices_p = 0
                high_mean_choices_num_p = 0
                equal_mean_choices_num_p = 0
                # split data
                data_partial[['choice_copy', 'reward']] = data_partial['outcome'].str.split(':', n=1).apply(pd.Series)  # split reward
                data_partial = data_partial.reset_index(drop=True)
                options = data_partial['choice'].unique()
                rewards = pd.DataFrame(columns=['value', 'counts', 'proportion'])
                data_partial_num = pd.DataFrame()
                high_ev_set = []
                for k in range(0, len(options)):
                    option = options[k]
                    # print(option)
                    data_option = data_partial[data_partial['choice'] == option]
                    # read option table
                    option_inf_sub = option_inf[option_inf['option'] == option]
                    outcome_values = option_inf_sub.filter(regex='^out').T.reset_index(drop=True).dropna()
                    outcome_proportions = option_inf_sub.filter(regex='^pr').T.reset_index(drop=True).dropna()
                    option_distribute = pd.concat([outcome_values, outcome_proportions], axis=1)
                    option_distribute.columns = ['value', 'proportion']
                    option_distribute = option_distribute.apply(pd.to_numeric, errors='coerce')
                    option_distribute = option_distribute.sort_values(by='value', ascending=True).reset_index(drop=True)
                    # consistent_outcomes
                    try:
                        set1 = set(data_option['reward'].astype(float))
                        set2 = set(option_distribute['value'].astype(float))
                    except Exception as e:
                        print(f"{study},option{option}: error-{e}")
                        continue
                    if set1.issubset(set2):
                        consistent_outcomes_p += 1
                    else:
                        print(set1, set2)
                        data_option = data_option[data_option['reward'].astype(float).isin(set2)]
                        # print(len(data_option))
                    data_partial_num = pd.concat([data_partial_num, data_option])
                    # consistent_probabilities
                    rews = data_option['reward'].value_counts().reset_index()
                    rews.columns = ['value', 'counts']
                    rews['proportion'] = rews['counts'] / sum(rews['counts'])
                    rews = rews.apply(pd.to_numeric, errors='coerce')
                    rews = rews.sort_values(by='value', ascending=True)
                    try:
                        chi2_cp, p_value_cp = chisquare(rews['proportion'], f_exp=option_distribute['proportion'])
                        dof_cp = len(rews['value']) - 1
                        if dof_cp == 0:
                            dof_cp += 1
                        consistent_probabilities_p.append(chi2_cp/dof_cp)
                        # consistent_probability_problem
                        if p_value_cp <= alpha:
                            consistent_probabilities_problem_p += 1
                    except Exception as e:
                        print(f"{study},option{option}: error-{e}")
                        continue
                    # high ev choices
                    ev = sum([x * y for x, y in zip(option_distribute['value'], option_distribute['proportion'])])
                    high_ev_set.append(ev)
                    if k == 0 or ev > high_ev_option_value_p:
                        high_ev_option_value_p = ev
                        high_ev_option_p = option
                # consistent outcome problem
                if consistent_outcomes_p == len(options):
                    consistent_outcomes_num += 1
                # consistent probabilities, for problem
                if len(consistent_probabilities_p) == 0:
                    consistent_probabilities_p.append(0)
                consistent_probabilities.append(sum(consistent_probabilities_p)/len(consistent_probabilities_p))
                # consistent probabilities problem
                if consistent_probabilities_problem_p == len(options):
                    consistent_probabilities_problem += 1
                # high ev choices
                data_partial_num = data_partial_num.reset_index(drop=True)
                if data_partial_num.empty:
                    print('data_partial_clean empty')
                    continue
                if high_ev_set.count(high_ev_option_value_p) > 1:
                    high_ev_choices_p = np.nan
                    high_ev_choices_problems_num.append(np.nan)
                else:
                    high_ev_trial_num += len(data_partial_num)
                    high_ev_choices_p = len(data_partial_num[data_partial_num['choice'] == high_ev_option_p])
                    high_ev_choices_test = binomtest(high_ev_choices_p, len(data_partial_num), p=1/option_number,
                                                     alternative='greater')
                    if high_ev_choices_test.pvalue <= alpha:
                        high_ev_choices_problems_num.append(1)
                    else:
                        high_ev_choices_problems_num.append(0)
                high_ev_choices_num.append(high_ev_choices_p)
                high_ev_choices_problem_prob.append(high_ev_choices_p/len(data_partial_num))

                # high mean choices
                participants = data_partial_num['subject'].unique()
                ignore_trial = 0
                for partiaipant in participants:
                    data_partial_num_participant = data_partial_num[data_partial_num['subject'] == partiaipant].sort_values('trial'). reset_index(drop=True)
                    for i in range(1, len(data_partial_num_participant)):
                        data_experienced = data_partial_num_participant.iloc[0:i].reset_index(drop=True)
                        if len(data_experienced['choice_copy'].unique()) < 2: # only high_experienced mean after experienced both options at least once
                            continue
                        else:
                            data_experienced['reward'] = data_experienced['reward'].astype(float)
                            reward_mean = data_experienced.groupby('choice')['reward'].mean()
                            reward_mean = reward_mean.astype(float)
                            max_reward_mean = reward_mean.idxmax()
                            if len(reward_mean[reward_mean == reward_mean[max_reward_mean]]) > 1:
                                equal_mean_choices_num_p += 1
                            elif data_partial_num_participant['choice'].iloc[i] == max_reward_mean:
                                high_mean_choices_num_p += 1
                high_mean_choices_num.append(high_mean_choices_num_p)
                equal_mean_choices_num.append(equal_mean_choices_num_p)
                # calculate high_mean_choices_problem_prob
                den_high = len(data_partial_num) - equal_mean_choices_num_p - ignore_trial - len(participants)

                if den_high == 0:
                    print("Error: divided by zero in high_mean_choices_problem_prob")
                    high_mean_choices_problem_prob.append(np.nan)
                    high_mean_choices_problems_num.append(np.nan)
                else:
                    high_mean_choices_problem_prob.append(high_mean_choices_num_p / den_high)
                    high_mean_choices_test = binomtest(high_mean_choices_num_p, den_high, p=1 / option_number,
                                                       alternative='greater')
                    if high_mean_choices_test.pvalue <= alpha:
                        high_mean_choices_problems_num.append(1)
                    else:
                        high_mean_choices_problems_num.append(0)

                # calculate equal_mean_choices_problem_prob
                den_equal = len(data_partial_num) - ignore_trial - len(participants)

                if den_equal == 0:
                    print("Error: divided by zero in equal_mean_choices_problem_prob")
                    equal_mean_choices_problem_prob.append(np.nan)
                else:
                    equal_mean_choices_problem_prob.append(equal_mean_choices_num_p / den_equal)
                # add cleaned data to all dataset
                cleaned_data = pd.concat([cleaned_data, data_partial_num])
                ignore_trials_num += ignore_trial
                participants_num += len(participants)

            # full feedback
            data_full = data_problem[data_problem['outcome'].str.contains('_', case=False)]
            if not data_full.empty:
                # validation_variables each problem
                consistent_outcomes_p = 0
                consistent_probabilities_p = []
                consistent_probabilities_problem_p = 0
                high_ev_option_value_p = 0
                high_ev_option_p = 0
                high_ev_choices_p = 0
                high_mean_choices_num_p = 0
                equal_mean_choices_num_p = 0
                # split data
                split_list = ['outcome1', 'outcome2', 'outcome3', 'outcome4', 'outcome5']
                max_outcomes = data_full['outcome'].str.count('_').max()
                data_full[split_list[0:max_outcomes+1]] = data_full['outcome'].str.split('_', expand=True).apply(pd.Series)
                choice_list = []
                reward_list = []
                for out in split_list[0:max_outcomes+1]:
                    ch = 'choice' + out[-1:]
                    re = 'reward' + out[-1:]
                    choice_list.append(ch)
                    reward_list.append(re)
                    data_full[[ch, re]] = data_full[out].str.split(':', expand=True).apply(pd.Series)
                data_full = data_full.reset_index(drop=True)
                # merge all outcome to one column
                choice_merge = pd.DataFrame(columns=['choice', 'reward'])
                for i in range(0, len(choice_list)):
                    choice_merge = pd.concat([choice_merge[['choice', 'reward']],
                                              data_full[[choice_list[i], reward_list[i]]].rename(
                                                  columns={choice_list[i]: 'choice', reward_list[i]: 'reward'})])
                choice_merge = choice_merge.reset_index(drop=True)
                options = data_full['choice'].unique()
                rewards = pd.DataFrame(columns=['value', 'counts', 'proportion'])
                data_full_num = data_full
                high_ev_set = []
                for k in range(0, len(options)):
                    option = options[k]
                    # print(option)
                    choice_merge_option = choice_merge[choice_merge['choice'] == option]
                    # read option table
                    option_inf_sub = option_inf[option_inf['option'] == option]
                    outcome_values = option_inf_sub.filter(regex='^out').T.reset_index(drop=True).dropna()
                    outcome_proportions = option_inf_sub.filter(regex='^pr').T.reset_index(drop=True).dropna()
                    option_distribute = pd.concat([outcome_values, outcome_proportions], axis=1)
                    option_distribute.columns = ['value', 'proportion']
                    option_distribute = option_distribute.apply(pd.to_numeric, errors='coerce')
                    option_distribute = option_distribute.sort_values(by='value', ascending=True).reset_index(drop=True)
                    # consistent_outcomes
                    try:
                        set1 = set(choice_merge_option['reward'].astype(float))
                        set2 = set(option_distribute['value'].astype(float))
                    except Exception as e:
                        print(f"{study},option{option}: error-{e}")
                        continue
                    if set1.issubset(set2):
                        consistent_outcomes_p += 1
                    else:
                        print(set1, set2)
                        choice_merge_option = choice_merge_option[choice_merge_option['reward'].astype(float).isin(set2)]
                        data_full_option = pd.DataFrame()
                        # for every choice i column, filter rows with unexpected values
                        for a in range(0, len(choice_list)):
                            data_a = data_full_num[data_full_num[choice_list[a]] == option]
                            data_a_clean = data_a[data_a[reward_list[a]].isin(set2)]
                            data_full_option = pd.concat([data_full_option, data_a_clean])
                        data_full_num = data_full_option
                    # consistent_probabilities
                    rews = choice_merge_option['reward'].value_counts().reset_index()
                    rews.columns = ['value', 'counts']
                    rews['proportion'] = rews['counts'] / sum(rews['counts'])
                    rews = rews.apply(pd.to_numeric, errors='coerce')
                    rews = rews.sort_values(by='value')
                    try:
                        chi2_cp, p_value_cp = chisquare(rews['proportion'], f_exp=option_distribute['proportion'])
                        dof_cp = len(rews['value']) - 1
                        if dof_cp == 0:
                            dof_cp += 1
                        consistent_probabilities_p.append(chi2_cp / dof_cp)
                        # consistent_probability_problem
                        if p_value_cp <= alpha:
                            consistent_probabilities_problem_p += 1
                    except Exception as e:
                        print(f"{study},option{option}: error-{e}")
                        continue
                    # high ev choices
                    ev = sum([x * y for x, y in zip(option_distribute['value'], option_distribute['proportion'])])
                    high_ev_set.append(ev)
                    if k == 0 or ev > high_ev_option_value_p:
                        high_ev_option_value_p = ev
                        high_ev_option_p = option
                # consistent outcome problem
                if consistent_outcomes_p == len(options):
                    consistent_outcomes_num += 1
                # consistent probabilities, for problem
                if len(consistent_probabilities_p) == 0:
                    consistent_probabilities_p.append(0)
                consistent_probabilities.append(sum(consistent_probabilities_p) / len(consistent_probabilities_p))
                # consistent probabilities problem
                if consistent_probabilities_problem_p == len(options):
                    consistent_probabilities_problem += 1
                # high ev choices
                data_full_num = data_full_num.reset_index(drop=True)
                if data_full_num.empty:
                    print('data_full_clean empty')
                    continue
                if high_ev_set.count(high_ev_option_value_p) > 1:
                    high_ev_choices_p = np.nan
                    high_ev_choices_problems_num.append(np.nan)
                else:
                    # print(high_ev_option_p)
                    high_ev_trial_num += len(data_full_num)
                    high_ev_choices_p = len(data_full_num[data_full_num['choice'] == high_ev_option_p])
                    high_ev_choices_test = binomtest(high_ev_choices_p, len(data_full_num), p=1/option_number,
                                                     alternative='greater')
                    if high_ev_choices_test.pvalue <= alpha:
                        high_ev_choices_problems_num.append(1)
                    else:
                        high_ev_choices_problems_num.append(0)
                high_ev_choices_num.append(high_ev_choices_p)
                high_ev_choices_problem_prob.append(high_ev_choices_p/len(data_full_num))
                # high mean choices
                participants = data_full_num['subject'].unique()
                for partiaipant in participants:
                    data_full_num_participant = data_full_num[data_full_num['subject'] == partiaipant].sort_values('trial').reset_index(drop=True)
                    for i in range(1, len(data_full_num_participant)):
                        data_experienced = data_full_num_participant.iloc[0:i].reset_index(drop=True)
                        df_merge = pd.DataFrame(columns=['choice', 'reward'])
                        for j in range(0, len(choice_list)):
                            df_merge = pd.concat([df_merge[['choice', 'reward']], data_experienced[[choice_list[j], reward_list[j]]].rename(columns={choice_list[j]: 'choice', reward_list[j]: 'reward'})])
                        df_merge = df_merge.reset_index()
                        df_merge['reward'] = df_merge['reward'].astype(float)
                        reward_mean = df_merge.groupby('choice')['reward'].mean()
                        reward_mean = reward_mean.astype(float)
                        max_reward_mean = reward_mean.idxmax()
                        if len(reward_mean[reward_mean == reward_mean[max_reward_mean]]) > 1:
                            equal_mean_choices_num_p += 1
                        elif data_full_num_participant['choice'][i] == max_reward_mean:
                            high_mean_choices_num_p += 1
                high_mean_choices_num.append(high_mean_choices_num_p)
                equal_mean_choices_num.append(equal_mean_choices_num_p)
                equal_mean_choices_problem_prob.append(equal_mean_choices_num_p/(len(data_full_num)-len(participants)))
                if len(data_full_num) - equal_mean_choices_num_p - len(participants) == 0:
                    high_mean_choices_problem_prob.append(np.nan)
                    high_mean_choices_problems_num.append(np.nan)
                else:
                    high_mean_choices_problem_prob.append(
                        high_mean_choices_num_p / (len(data_full_num) - equal_mean_choices_num_p - len(participants)))
                    high_mean_choices_test = binomtest(high_mean_choices_num_p, (
                                len(data_full_num) - equal_mean_choices_num_p - len(participants)), p=1 / option_number,
                                                       alternative='greater')
                    if high_mean_choices_test.pvalue <= alpha:
                        high_mean_choices_problems_num.append(1)
                    else:
                        high_mean_choices_problems_num.append(0)
                # add cleaned data to all dataset
                cleaned_data = pd.concat([cleaned_data, data_full_num])
                participants_num += len(participants)
            # is this problem has both partial and full feedback type
            if (not data_partial.empty) and (not data_full.empty):
                both = 2

        # write table for this study
        # if consistent_probabilities
        if len(consistent_probabilities) == 0:
            consistent_probabilities.append(0)
        if len(cleaned_data) == 0:
            print('no consist data')
            continue
        # for overall high ev choice proportion
        if high_ev_trial_num == 0:
            high_ev_choice = str(np.nan)
        else:
            high_ev_choice = round(np.nansum(high_ev_choices_num) / high_ev_trial_num, 3)
        # for high ev choice proportion for each problem
        problem_with_high_ev = [x for x in high_ev_choices_problems_num if not math.isnan(x)]
        if len(problem_with_high_ev) == 0:
            high_ev_proportion = str(np.nan)
        else:
            high_ev_proportion = round(np.nansum(high_ev_choices_problems_num) / len(problem_with_high_ev), 3)
        # for high mean choice proportion for each problem
        problem_with_high_mean = [x for x in high_mean_choices_problems_num if not math.isnan(x)]
        if len(problem_with_high_mean) == 0:
            high_mean_proportion = str(np.nan)
        else:
            high_mean_proportion = round(np.nansum(high_mean_choices_problems_num)/len(problem_with_high_mean), 3)
        # print(consistent_outcomes_num, problem_list, both, consistent_outcomes_num / (len(problem_list) * both))
        new_row = [study, round(consistent_outcomes_num / (len(option_pairs) * both), 3), round(sum(consistent_probabilities)/len(consistent_probabilities), 3),
                   round(consistent_probabilities_problem / (len(option_pairs) * both), 3), high_ev_choice,
                   high_ev_proportion, ','.join(map(str, high_ev_choices_problems_num)),
                   ','.join(map(str, [round(num, 3) for num in high_ev_choices_problem_prob])),
                   round(np.nansum(high_mean_choices_num) / (len(cleaned_data)-len(equal_mean_choices_num)-ignore_trials_num - participants_num), 3),
                   high_mean_proportion, ','.join(map(str, high_mean_choices_problems_num)),
                   ','.join(map(str, [round(num, 3) for num in high_mean_choices_problem_prob])),
                   round(np.nansum(equal_mean_choices_num)/(len(cleaned_data)-ignore_trials_num - participants_num), 3),
                   ','.join(map(str, [round(num, 3) for num in equal_mean_choices_problem_prob]))]
        validation_df.loc[len(validation_df)] = new_row
        wb = load_workbook(save_path)
        ws = wb.active
        ws.append(new_row)
        wb.save(save_path)
        # end signal
        current_time = datetime.datetime.now()
        print(study, current_time)
